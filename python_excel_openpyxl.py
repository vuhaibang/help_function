import openpyxl

url = 'https://realpython.com/openpyxl-excel-spreadsheets-python/'

wb = openpyxl.load_workbook('LENDBOX.xlsx')

# Select sheet by index: from 0 to end
sheet1 = wb.worksheets[0]
# Select sheet by name:
sheet2 = wb.get_sheet_names('name')

# Select range
cell_range = sheet1['A1':'C2']

# Select cells
for column in range(1, 200):
    for row in range(1, 33):
        a = sheet2.cell(row=row, column=column).value


# Font
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

font = Font(name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')
fill = PatternFill(fill_type=None,
                   start_color='FFFFFFFF',
                   end_color='FF000000')
border = Border(left=Side(border_style=None,
                          color='FF000000'),
                right=Side(border_style=None,
                           color='FF000000'),
                top=Side(border_style=None,
                         color='FF000000'),
                bottom=Side(border_style=None,
                            color='FF000000'),
                diagonal=Side(border_style=None,
                              color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style=None,
                             color='FF000000'),
                vertical=Side(border_style=None,
                              color='FF000000'),
                horizontal=Side(border_style=None,
                                color='FF000000')
                )
alignment = Alignment(horizontal='general',
                      vertical='bottom',
                      text_rotation=0,
                      wrap_text=False,
                      shrink_to_fit=False,
                      indent=0)
number_format = 'General'
protection = Protection(locked=True,
                        hidden=False)

# Save
wb.save('name.xlsx')
